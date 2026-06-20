"""Export scored leads to CSV and a formatted Excel workbook."""
from __future__ import annotations
import csv
from .models import Lead

COLUMNS = [
    ("lead_score", "Lead Score"), ("priority", "Priority"),
    ("name", "Name"), ("role", "Role"), ("company", "Company"),
    ("event_type", "Event Type"), ("event_date", "Event Date"),
    ("liquidity_signal", "Liquidity Signal"),
    ("est_liquid_wealth_usd_m", "Est. Liquid Wealth ($m)"),
    ("event_value_usd_m", "Event Value ($m)"), ("stake_pct", "Stake %"),
    ("exchange_ticker", "Exchange / Ticker"),
    ("diaspora_category", "Diaspora Category"),
    ("city", "City"), ("country", "Country"), ("region", "Region"),
    ("industry", "Industry"),
    ("recency_score", "Recency"), ("event_score", "Magnitude"),
    ("addressability", "Addressability"), ("diaspora_fit", "Diaspora Fit"), ("geo_fit", "Geo Fit"),
    ("public_source", "Public Source"), ("source_url", "Source URL"), ("notes", "Notes"),
]


def to_csv(leads: list[Lead], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([h for _, h in COLUMNS])
        for lead in leads:
            d = lead.to_dict()
            w.writerow([d.get(k, "") for k, _ in COLUMNS])


def to_excel(leads: list[Lead], path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl not installed. Run: pip install -r requirements.txt")
    wb = Workbook(); ws = wb.active; ws.title = "Leads"
    headers = [h for _, h in COLUMNS]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F3864")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for lead in leads:
        d = lead.to_dict()
        ws.append([d.get(k, "") for k, _ in COLUMNS])
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(leads)+1}"
    wb.save(path)
