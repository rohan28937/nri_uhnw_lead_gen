"""Builds the EXCLUSIVE, signal-driven seed dataset + company watchlist.

Focus: non-obvious, newly-liquid Indian-origin (NRI/OCI) and resident-India founders/
execs -- NOT the saturated billionaire rich list. Every row is anchored to a real,
public liquidity event (recent IPO, acquisition, large funding round).

$ amounts in MILLIONS. Stakes & event facts sourced from filings/news (source_url).
est_liquid_wealth is a rough MODELED figure for prioritisation only -- VERIFY before
outreach. See COMPLIANCE.md.
"""
import csv, sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from nri_leadgen.models import Lead
from nri_leadgen.scoring import score_lead

R = [
 ["Bipul Sinha","Co-founder & CEO","Rubrik","OCI / Indian-origin","Palo Alto, CA","USA","USA & Canada","Cybersecurity / Data security","NYSE: RBRK",
  "ipo","2024-04",752,7.6,900,"IPO Apr-2024; ~7.6% founder stake now public & vesting","Rubrik S-1 / CNBC","https://www.cnbc.com/2024/04/01/rubrik-files-to-go-public.html",4,5,5],
 ["Arvind Nithrakashyap","Co-founder & CTO","Rubrik","OCI / Indian-origin","Palo Alto, CA","USA","USA & Canada","Cybersecurity / Data security","NYSE: RBRK",
  "ipo","2024-04",752,6.7,780,"IPO Apr-2024; ~6.7% founder stake","Rubrik S-1 / Blocks&Files","https://blocksandfiles.com/2024/04/15/rubrik-ipo-to-raise-up-to-713-million/",4,5,5],
 ["Soham Mazumdar","Co-founder","Rubrik","OCI / Indian-origin","San Francisco, CA","USA","USA & Canada","Cybersecurity","NYSE: RBRK",
  "ipo","2024-04",752,None,300,"IPO Apr-2024; founding engineer equity","Rubrik S-1","https://en.wikipedia.org/wiki/Rubrik",3,5,5],
 ["Arvind Jain","Co-founder (Rubrik) & Founder/CEO (Glean)","Glean / Rubrik","OCI / Indian-origin","Palo Alto, CA","USA","USA & Canada","AI / Enterprise search","Private (Glean) + NYSE: RBRK",
  "funding_round","2025-06",260,7.0,1000,"DOUBLE signal: ~7% Rubrik stake (now public) + Glean AI unicorn mega-round","TechCrunch / Rubrik S-1","https://techcrunch.com/2024/04/25/rubriks-shares-climb-20-in-its-public-debut",4,5,5],
 ["Jitendra Mohan","Co-founder & CEO","Astera Labs","OCI / Indian-origin (verify)","Santa Clara, CA","USA","USA & Canada","Semiconductors / AI connectivity","NASDAQ: ALAB",
  "ipo","2024-03",713,None,1200,"IPO Mar-2024 at $5.5B; stock re-rated sharply post-IPO","Astera S-1 / INO","https://www.ino.com/blog/2024/03/intel-backed-astera-labs-goes-public/",4,4,5],
 ["Sanjay Gajendra","Co-founder, President & COO","Astera Labs","OCI / Indian-origin (verify)","Santa Clara, CA","USA","USA & Canada","Semiconductors / AI connectivity","NASDAQ: ALAB",
  "ipo","2024-03",713,None,800,"IPO Mar-2024; co-founder equity now public","Astera S-1","https://en.wikipedia.org/wiki/Astera_Labs",4,4,5],
 ["Sunny Madra","President (Groq)","Groq / NVIDIA","OCI / Indian-origin (verify)","Mountain View, CA","USA","USA & Canada","AI chips","Private (NVIDIA deal)",
  "acquisition","2025-12",20000,None,300,"NVIDIA $20B licensing/asset deal acqui-hired Groq leadership Dec-2025","Tech Startups","https://techstartups.com/2025/12/25/nvidia-strikes-20b-licensing-and-asset-deal-with-ai-chip-startup-groq/",3,3,5],
 ["Sridhar Ramaswamy","CEO (ex-founder, Neeva)","Snowflake","OCI / Indian-origin","Bay Area, CA","USA","USA & Canada","Data cloud / AI","NYSE: SNOW",
  "new_insider","2024-02",None,None,150,"Became Snowflake CEO Feb-2024 after Neeva sale; large RSU + Neeva proceeds","news","https://www.snowflake.com",4,5,5],
 ["Dev Ittycheria","President & CEO","MongoDB","OCI / Indian-origin","New York, NY","USA","USA & Canada","Database / Cloud","NASDAQ: MDB",
  "promoter_holding","2024-01",None,None,200,"Long-tenured CEO; recurring Form 4 sales = realised liquidity","SEC EDGAR","https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&company=mongodb",4,5,5],
 ["Rajiv Ramaswami","President & CEO","Nutanix","OCI / Indian-origin","San Jose, CA","USA","USA & Canada","Cloud infrastructure","NASDAQ: NTNX",
  "promoter_holding","2024-01",None,None,90,"CEO since 2020; regular insider transactions","SEC EDGAR","https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&company=nutanix",4,5,5],
 ["Ravi Kumar S","CEO","Cognizant","OCI / Indian-origin","Teaneck, NJ","USA","USA & Canada","IT services","NASDAQ: CTSH",
  "promoter_holding","2023-01",None,None,60,"CEO since Jan-2023; equity comp + Infosys-era wealth","SEC EDGAR","https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&company=cognizant",4,5,5],
 ["Keshav Murugesh","Group CEO","WNS Holdings","OCI / Indian-origin","Mumbai / NY","India","India","BPM / IT services","NYSE: WNS",
  "promoter_holding","2023-01",None,None,80,"Long-tenured CEO of NYSE-listed WNS; disclosed holdings","SEC EDGAR","https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&company=wns",3,5,4],
 ["Peyush Bansal","Co-founder & CEO","Lenskart","Resident India","Gurugram","India","India","Consumer / Retail tech","NSE/BSE: LENSKART",
  "ipo","2025-10",None,None,500,"2025 IPO; founder stake now listed & partially monetised","Inc42 IPO Tracker","https://inc42.com/features/indian-startup-ipo-tracker-2026/",3,5,5],
 ["Lalit Keshre","Co-founder & CEO","Groww","Resident India","Bengaluru","India","India","Fintech / Broking","NSE/BSE: GROWW",
  "ipo","2025-11",None,None,400,"2025 IPO; founder equity listed","Inc42 IPO Tracker","https://inc42.com/features/indian-startup-ipo-tracker-2026/",3,5,5],
 ["Tarun Mehta","Co-founder & CEO","Ather Energy","Resident India","Bengaluru","India","India","EV / Mobility","NSE/BSE: ATHER",
  "ipo","2025-05",None,None,250,"2025 IPO; founder stake listed","Inc42 IPO Tracker","https://inc42.com/features/indian-startup-ipo-tracker-2026/",3,5,5],
 ["Alakh Pandey","Co-founder & CEO","PhysicsWallah","Resident India","Noida","India","India","Edtech","NSE/BSE: PW",
  "ipo","2025-11",None,None,300,"2025 IPO; founder-led edtech, fresh public equity","Inc42 IPO Tracker","https://inc42.com/features/indian-startup-ipo-tracker-2026/",3,5,5],
 ["Abhiraj Singh Bhal","Co-founder & CEO","Urban Company","Resident India","Gurugram","India","India","Consumer services","NSE/BSE: URBANCO",
  "ipo","2025-09",None,None,200,"2025 IPO; co-founder equity now public","Inc42 IPO Tracker","https://inc42.com/features/indian-startup-ipo-tracker-2026/",3,5,5],
 ["Vidit Aatrey","Co-founder & CEO","Meesho","Resident India","Bengaluru","India","India","E-commerce","NSE/BSE: MEESHO",
  "ipo","2025-12",None,None,300,"2025 IPO; founder stake listed","Inc42 IPO Tracker","https://inc42.com/features/indian-startup-ipo-tracker-2026/",3,5,5],
 ["LuLu Retail insiders (post-IPO)","Founders/execs","LuLu Retail Holdings","NRI","Abu Dhabi","UAE","Gulf","Retail","ADX: LULU",
  "ipo","2024-11",1430,None,200,"Nov-2024 ADX IPO; senior NRI execs hold vested equity (enrich individuals)","ADX / news","https://www.adx.ae",3,4,5],
]

HEADERS = ["Lead Score","Priority","Name","Role","Company","Event Type","Event Date",
 "Liquidity Signal","Est. Liquid Wealth ($m)","Event Value ($m)","Stake %","Exchange / Ticker",
 "Diaspora Category","City","Country","Region","Industry","Recency","Magnitude",
 "Addressability","Diaspora Fit","Geo Fit","Public Source","Source URL","Notes"]

with open("data/seed_leads.csv","w",newline="",encoding="utf-8") as f:
    w=csv.writer(f)
    w.writerow(["name","role","company","diaspora_category","city","country","region","industry",
                "exchange_ticker","event_type","event_date","event_value_usd_m","stake_pct",
                "est_liquid_wealth_usd_m","liquidity_signal","public_source","source_url",
                "addressability","diaspora_fit","geo_fit"])
    for r in R:
        w.writerow(["" if x is None else x for x in r])

leads=[]
for r in R:
    (name,role,company,dia,city,country,region,industry,ticker,etype,edate,eval_m,stake,
     est_m,signal,source,url,addr,dfit,gfit)=r
    leads.append(score_lead(Lead(
        name=name,role=role,company=company,diaspora_category=dia,city=city,country=country,
        region=region,industry=industry,exchange_ticker=ticker,event_type=etype,event_date=edate,
        event_value_usd_m=eval_m,stake_pct=stake,est_liquid_wealth_usd_m=est_m,liquidity_signal=signal,
        public_source=source,source_url=url,addressability=addr,diaspora_fit=dfit,geo_fit=gfit)))
leads.sort(key=lambda l:l.lead_score,reverse=True)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
wb=Workbook(); ws=wb.active; ws.title="Exclusive Leads"
FONT="Arial"
hf=PatternFill("solid",fgColor="1F3864"); hfont=Font(name=FONT,bold=True,color="FFFFFF",size=10)
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
ahot=PatternFill("solid",fgColor="C6EFCE"); awarm=PatternFill("solid",fgColor="FFEB9C")
ws.append(HEADERS)
for c in range(1,len(HEADERS)+1):
    cell=ws.cell(row=1,column=c); cell.fill=hf; cell.font=hfont; cell.border=border
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
for l in leads:
    d=l.to_dict()
    ws.append([l.lead_score,l.priority,d["name"],d["role"],d["company"],d["event_type"],d["event_date"],
        d["liquidity_signal"],d["est_liquid_wealth_usd_m"],d["event_value_usd_m"],d["stake_pct"],
        d["exchange_ticker"],d["diaspora_category"],d["city"],d["country"],d["region"],d["industry"],
        l.recency_score,l.event_score,d["addressability"],d["diaspora_fit"],d["geo_fit"],
        d["public_source"],d["source_url"],d["notes"]])
    row=ws.max_row
    band=ws.cell(row=row,column=2).value
    fillc = ahot if band=="A - Hot" else awarm if band=="B - Warm" else None
    for c in range(1,len(HEADERS)+1):
        cell=ws.cell(row=row,column=c); cell.border=border
        cell.font=Font(name=FONT,size=10)
        cell.alignment=Alignment(vertical="center",wrap_text=(c in (8,25)))
        if c<=2 and fillc: cell.fill=fillc
W=[8,11,20,26,18,14,11,38,14,13,8,22,20,14,10,14,20,8,9,12,11,8,18,30,40]
for i,w in enumerate(W,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="C2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(HEADERS))}{len(leads)+1}"

ws2=wb.create_sheet("README")
for r in [
 ["Exclusive UHNW NRI/OCI Leads - signal-driven (NOT a billionaire rich list)"],[""],
 ["Thesis","Targets non-obvious, newly-liquid Indian-origin founders/execs/ESOP holders. The unit of value is a LIQUIDITY EVENT (recent IPO, acquisition, big funding/secondary), not rich-list rank."],
 ["Score (0-100)","Event recency 30% + Event magnitude 25% + Addressability 20% + Diaspora fit 15% + Geo fit 10%. Recency is date-relative, so re-run the pipeline to refresh."],
 ["Bands","A-Hot >=80 (green), B-Warm 65-79 (amber), C-Develop <65."],
 ["$ units","All money in $ MILLIONS. 'Est. Liquid Wealth' is a MODELED prioritisation estimate, not a verified figure."],
 ["How to scale","Run the pipeline with SEC EDGAR enabled + the company watchlist (data/watchlist.csv): it expands each ticker into ALL its insiders and flags recent Form 4 SALES - that surfaces the CFOs/VPs/board members no rich list covers."],
 ["Verify","Confirm identity, Indian origin/residency and current holdings before any outreach. See COMPLIANCE.md."],
]:
    ws2.append(r)
ws2.column_dimensions["A"].width=20; ws2.column_dimensions["B"].width=110
ws2["A1"].font=Font(name=FONT,bold=True,size=13,color="1F3864")
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment=Alignment(vertical="top",wrap_text=True)
        cell.font=Font(name=FONT,bold=(cell.column==1),size=10)
wb.save("data/Exclusive_NRI_OCI_Leads.xlsx")

WATCH=[
 ["RBRK","Rubrik","NYSE","Bipul Sinha / Arvind Nithrakashyap","Co-founders","IPO Apr-2024","Indian-origin founders; track lock-up sales"],
 ["ALAB","Astera Labs","NASDAQ","Jitendra Mohan / Sanjay Gajendra","Co-founders","IPO Mar-2024","AI connectivity; sharp post-IPO re-rate"],
 ["ZS","Zscaler","NASDAQ","Jay Chaudhry","Founder & CEO","Listed","Track CFO/VP insiders too, not just founder"],
 ["ANET","Arista Networks","NYSE","Jayshree Ullal","Chair (ex-CEO)","Listed","Long-tenured insiders with recurring sales"],
 ["SNOW","Snowflake","NYSE","Sridhar Ramaswamy","CEO","CEO since 2024","Ex-Neeva; new RSU grants"],
 ["MDB","MongoDB","NASDAQ","Dev Ittycheria","CEO","Listed","Regular Form 4 sales"],
 ["NTNX","Nutanix","NASDAQ","Rajiv Ramaswami","CEO","Listed","Insider transactions"],
 ["PANW","Palo Alto Networks","NASDAQ","Nikesh Arora","Chairman & CEO","Listed","Very large RSU comp"],
 ["WDAY","Workday","NASDAQ","Aneel Bhusri","Co-founder, Chair","Listed","Founder stake"],
 ["W","Wayfair","NYSE","Niraj Shah","Co-founder & CEO","Listed","Founder stake"],
 ["CTSH","Cognizant","NASDAQ","Ravi Kumar S","CEO","CEO since 2023","Equity comp ramping"],
 ["WNS","WNS Holdings","NYSE","Keshav Murugesh","Group CEO","Listed","NYSE-listed Indian BPM"],
 ["FRSH","Freshworks","NASDAQ","Girish Mathrubootham","Founder, Chair","IPO 2021","First Indian SaaS on Nasdaq"],
]
with open("data/watchlist.csv","w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(["ticker","company","exchange","key_indian_origin_leader","role","event","note"])
    w.writerows(WATCH)

print(f"Wrote {len(leads)} exclusive leads -> Exclusive_NRI_OCI_Leads.xlsx, seed_leads.csv, watchlist.csv ({len(WATCH)} tickers)")
